import sys
import openpyxl
from openpyxl.styles import PatternFill
from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5.QtGui import *
from data import *
form_class = uic.loadUiType("show.ui")[0]
class MyWindow(QMainWindow, form_class):
    excelDatalist = ExcelDatalist()
    updown = 0
    fname =''
    wb=''

    def messageBox(self):
        msgBox = QMessageBox()
        msgBox.setWindowTitle("저장")  # 메세지창의 상단 제목
        msgBox.setIcon(QMessageBox.Information)  # 메세지창 내부에 표시될 아이콘
        msgBox.setText("저장하시겠습니까?")  # 메세지 제목
        # msgBox.setInformativeText("저장하시겠습니까?")          # 메세지 내용
        msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)  # 메세지창의 버튼
        msgBox.setDefaultButton(QMessageBox.Yes)  # 포커스가 지정된 기본 버튼
        return msgBox.exec_()

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.sortButton.clicked.connect(self.btn_clicked)
        self.actionfile_open.triggered.connect(self.open_file_Clicked)
        self.actionfile_save.triggered.connect(self.save_file_Clicked)

    def save_file_Clicked(self):

        if self.updown == 0:
            QMessageBox.about(self, "message", "파일을 먼저 로딩해주세요")
        else:
            reslut =self.messageBox()
            if reslut == QMessageBox.Yes:
                ws = self.wb.active
                dlist = self.excelDatalist.datalist
                yellowFill = PatternFill(start_color='FFFFFF00',
                                         end_color='FFFFFF00',
                                         fill_type='solid')
                whiteFill = PatternFill(start_color='FFFFFF',
                                         end_color='FFFFFF',
                                         fill_type='solid')
                index = 0
                for r in ws.rows:
                    if r[0].row > 2:
                        r[0].value = dlist[index].name
                        r[1].value = dlist[index].ratio
                        r[2].value = str(dlist[index].ratio) + "%"
                        if r[0].value == "경북대학교":
                            r[0].fill = yellowFill
                            r[1].fill = yellowFill
                            r[2].fill = yellowFill
                        else:
                            r[0].fill = whiteFill
                            r[1].fill = whiteFill
                            r[2].fill = whiteFill

                        index = index + 1
                self.wb.save(self.fname[0])



    def btn_clicked(self):
        if self.updown==1:
            self.excelDatalist.sort_data_up()
            self.updown =2
            self.dataTable.clearContents()
            self.set_data_inTable()
            #self.excelDatalist.printlist()

        elif self.updown==2:
            self.excelDatalist.sort_data_down()
            self.updown = 1
            self.dataTable.clearContents()
            self.set_data_inTable()
            #self.excelDatalist.printlist()
        else:
            QMessageBox.about(self, "message", "파일을 먼저 로딩해주세요")


    def open_file_Clicked(self):
        self.fname = QFileDialog.getOpenFileName(self)
        self.wb = openpyxl.load_workbook(self.fname[0])
        ws = self.wb.active
        for r in ws.rows:
            if r[0].row > 2:
                row_name = r[0].value
                row_ratio = float(r[1].value)
                data = ExcelData(row_name,row_ratio)
                self.excelDatalist.push(data)
        #self.excelDatalist.printlist()

        self.updown =1
        self.dataTable.clearContents()
        size =len(self.excelDatalist.datalist)
        self.dataTable.setRowCount(size)
        self.set_data_inTable()


    def set_data_inTable(self):
        index = 0
        colorVar = QColor(255, 255, 0, 255)
        for data in self.excelDatalist.datalist:
            self.dataTable.setItem(index, 0, QTableWidgetItem(data.name))
            self.dataTable.setItem(index, 1, QTableWidgetItem(str(data.ratio)))
            self.dataTable.setItem(index, 2, QTableWidgetItem(str(data.ratio) + "%"))

            if data.name =="경북대학교":
                self.dataTable.item(index, 0).setBackground(colorVar)
                self.dataTable.item(index, 1).setBackground(colorVar)
                self.dataTable.item(index, 2).setBackground(colorVar)

            index = index + 1
        self.dataTable.resizeColumnsToContents()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = MyWindow()
    myWindow.show()
    app.exec_()